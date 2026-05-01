# from django.shortcuts import render, redirect, get_object_or_404
# from django.contrib.auth.decorators import login_required
# from django.contrib import messages
# from django.http import JsonResponse
# from .models import Timetable, BellSchedule, Period, Lesson, TimetableSlot, DAYS
# from resources.models import Subject, Grade, Division, Room
# from faculty.models import Faculty
# from .scheduler import generate_timetable

# @login_required
# def dashboard(request):
#     timetables = Timetable.objects.filter(owner=request.user)
#     faculty_count = Faculty.objects.filter(owner=request.user).count()
#     subject_count = Subject.objects.filter(owner=request.user).count()
#     grade_count = Grade.objects.filter(owner=request.user).count()
#     ctx = {
#         'timetables': timetables,
#         'faculty_count': faculty_count,
#         'subject_count': subject_count,
#         'grade_count': grade_count,
#         'published': timetables.filter(status='published').first(),
#     }
#     return render(request, 'dashboard/home.html', ctx)

# @login_required
# def timetable_list(request):
#     timetables = Timetable.objects.filter(owner=request.user)
#     return render(request, 'timetables/list.html', {'timetables': timetables})

# @login_required
# def create_timetable(request):
#     if request.method == 'POST':
#         name = request.POST.get('name', '').strip()
#         if not name:
#             messages.error(request, 'Timetable name is required.')
#             return redirect('timetable_list')
#         tt = Timetable.objects.create(
#             owner=request.user,
#             name=name,
#             academic_year=request.POST.get('academic_year', ''),
#         )
#         messages.success(request, f'Timetable "{name}" created.')
#         return redirect('timetable_setup', pk=tt.pk)
#     return redirect('timetable_list')

# @login_required
# def timetable_setup(request, pk):
#     tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
#     bell_schedules = BellSchedule.objects.filter(owner=request.user)
#     grades = Grade.objects.filter(owner=request.user).prefetch_related('divisions')
#     faculty_qs = Faculty.objects.filter(owner=request.user, status='active')
#     subjects = Subject.objects.filter(owner=request.user)
#     rooms = Room.objects.filter(owner=request.user)
#     lessons = tt.lessons.select_related('subject', 'faculty', 'division', 'room').all()

#     if request.method == 'POST':
#         action = request.POST.get('action')

#         if action == 'assign_bell':
#             bell_id = request.POST.get('bell_schedule')
#             if bell_id:
#                 tt.bell_schedule = BellSchedule.objects.get(pk=bell_id, owner=request.user)
#                 tt.save()
#                 messages.success(request, 'Bell schedule assigned.')

#         elif action == 'add_lesson':
#             subject_id = request.POST.get('subject')
#             faculty_id = request.POST.get('faculty')
#             division_id = request.POST.get('division')
#             ppw = request.POST.get('periods_per_week', 4)
#             room_id = request.POST.get('room')
#             if subject_id and division_id:
#                 Lesson.objects.create(
#                     timetable=tt,
#                     subject_id=subject_id,
#                     faculty_id=faculty_id if faculty_id else None,
#                     division_id=division_id,
#                     room_id=room_id if room_id else None,
#                     periods_per_week=ppw,
#                 )
#                 messages.success(request, 'Lesson added.')
#             else:
#                 messages.error(request, 'Subject and Division are required.')

#         elif action == 'delete_lesson':
#             lesson_id = request.POST.get('lesson_id')
#             Lesson.objects.filter(pk=lesson_id, timetable=tt).delete()
#             messages.success(request, 'Lesson removed.')

#         elif action == 'generate':
#             tt.status = 'generating'
#             tt.save()
#             success, log = generate_timetable(tt)
#             tt.generation_log = log
#             tt.ai_generated = True
#             tt.status = 'generated' if success else 'draft'
#             tt.save()
#             if success:
#                 messages.success(request, 'Timetable generated successfully!')
#             else:
#                 messages.error(request, f'Generation failed: {log[:200]}')
#             return redirect('timetable_view', pk=tt.pk)

#         elif action == 'publish':
#             tt.status = 'published'
#             tt.save()
#             messages.success(request, 'Timetable published!')

#         return redirect('timetable_setup', pk=pk)

#     ctx = {
#         'tt': tt,
#         'bell_schedules': bell_schedules,
#         'grades': grades,
#         'faculty': faculty_qs,
#         'subjects': subjects,
#         'rooms': rooms,
#         'lessons': lessons,
#         'days': DAYS,
#         'setup_steps': get_setup_steps(tt),
#     }
#     return render(request, 'timetables/setup.html', ctx)

# def get_setup_steps(tt):
#     return {
#         'bell': tt.bell_schedule is not None,
#         'lessons': tt.lessons.exists(),
#         'generated': tt.status in ['generated', 'published'],
#     }

# @login_required
# def timetable_view(request, pk):
#     tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
#     bell = tt.bell_schedule
#     periods = []
#     if bell:
#         periods = list(bell.periods.order_by('period_number'))
#     working_days = bell.working_days if bell else DAYS[:5]

#     # Build grid: {division_id: {day: {period_id: slot}}}
#     slots = tt.slots.select_related(
#         'lesson__subject', 'lesson__division', 'lesson__faculty', 'period', 'room'
#     ).all()

#     divisions = Division.objects.filter(
#         id__in=tt.lessons.values_list('division_id', flat=True)
#     ).select_related('grade')

#     grid = {}
#     for div in divisions:
#         grid[div.id] = {day: {} for day in working_days}

#     for slot in slots:
#         div_id = slot.lesson.division_id
#         if div_id in grid and slot.day in grid[div_id]:
#             grid[div_id][slot.day][slot.period_id] = slot

#     ctx = {
#         'tt': tt,
#         'periods': periods,
#         'working_days': working_days,
#         'divisions': divisions,
#         'grid': grid,
#         'slots': slots,
#     }
#     return render(request, 'timetables/view.html', ctx)

# @login_required
# def delete_timetable(request, pk):
#     tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
#     name = tt.name
#     tt.delete()
#     messages.success(request, f'Timetable "{name}" deleted.')
#     return redirect('timetable_list')

# # Bell Schedule Views
# @login_required
# def bell_list(request):
#     bells = BellSchedule.objects.filter(owner=request.user).prefetch_related('periods')
#     return render(request, 'timetables/bells.html', {'bells': bells})

# @login_required
# def create_bell(request):
#     if request.method == 'POST':
#         name = request.POST.get('name', '').strip()
#         days = request.POST.getlist('working_days')
#         if name and days:
#             bell = BellSchedule.objects.create(owner=request.user, name=name, working_days=days)
#             # Add periods
#             period_names = request.POST.getlist('period_name')
#             start_times = request.POST.getlist('start_time')
#             end_times = request.POST.getlist('end_time')
#             period_types = request.POST.getlist('period_type')
#             for i, pname in enumerate(period_names):
#                 if pname.strip() and i < len(start_times) and i < len(end_times):
#                     Period.objects.create(
#                         bell_schedule=bell,
#                         name=pname.strip(),
#                         period_number=i + 1,
#                         start_time=start_times[i],
#                         end_time=end_times[i],
#                         period_type=period_types[i] if i < len(period_types) else 'class',
#                     )
#             messages.success(request, f'Bell schedule "{name}" created.')
#         else:
#             messages.error(request, 'Name and working days are required.')
#     return redirect('bell_list')

# @login_required
# def delete_bell(request, pk):
#     bell = get_object_or_404(BellSchedule, pk=pk, owner=request.user)
#     bell.delete()
#     messages.success(request, 'Bell schedule deleted.')
#     return redirect('bell_list')



# ============================================================
# timetables/views.py  — COMPLETE FILE (replace entirely)
# ============================================================
import json
import io
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views import View


# ────────────────────────────────────────────────────────────
# TIMETABLE LIST
# ────────────────────────────────────────────────────────────
class TimetableListView(LoginRequiredMixin, ListView):
    template_name = 'timetables/list.html'
    context_object_name = 'timetables'

    def get_queryset(self):
        from timetables.models import Timetable
        return Timetable.objects.filter(owner=self.request.user)


# ────────────────────────────────────────────────────────────
# TIMETABLE CREATE
# ────────────────────────────────────────────────────────────
class TimetableCreateView(LoginRequiredMixin, View):
    def get(self, request):
        from timetables.forms import TimetableForm
        return render(request, 'timetables/create.html', {
            'form': TimetableForm(user=request.user),
        })

    def post(self, request):
        from timetables.forms import TimetableForm
        form = TimetableForm(request.POST, user=request.user)
        if form.is_valid():
            tt = form.save(commit=False)
            tt.owner = request.user
            tt.save()
            messages.success(request, f'Timetable "{tt.name}" created!')
            return redirect('timetables:setup', pk=tt.pk)
        return render(request, 'timetables/create.html', {'form': form})


# ────────────────────────────────────────────────────────────
# TIMETABLE SETUP
# ────────────────────────────────────────────────────────────
class TimetableSetupView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import Timetable
        from timetables.forms import LessonForm
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        return render(request, 'timetables/setup.html', self._ctx(request, tt, LessonForm(user=request.user)))

    def _ctx(self, request, tt, lesson_form):
        return {
            'timetable'   : tt,
            'bell_periods': tt.bell_schedule.periods.all().order_by('period_number') if tt.bell_schedule else [],
            'lessons'     : tt.lessons.select_related('division__grade', 'subject', 'faculty').all(),
            'lesson_form' : lesson_form,
            'has_bell'    : tt.bell_schedule is not None,
            'has_lessons' : tt.lessons.exists(),
            'bell_count'  : tt.bell_schedule.periods.filter(period_type='class').count() if tt.bell_schedule else 0,
            'lesson_count': tt.lessons.count(),
            'slot_count'  : tt.slots.count(),
        }


# ────────────────────────────────────────────────────────────
# BELL SCHEDULE LIST + CREATE
# ────────────────────────────────────────────────────────────
# class BellScheduleListView(LoginRequiredMixin, View):
#     def get(self, request):
#         from timetables.models import BellSchedule
#         bells = BellSchedule.objects.filter(owner=request.user).prefetch_related('periods')
#         return render(request, 'timetables/bells.html', {'bells': bells})

class BellScheduleDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import BellSchedule
        from timetables.forms import PeriodForm
        bell = get_object_or_404(BellSchedule, pk=pk, owner=request.user)
        return render(request, 'timetables/bell_detail.html', {
            'bell': bell,
            'periods': bell.periods.all().order_by('period_number'),
            'period_form': PeriodForm(),
        })

class BellScheduleCreateView(LoginRequiredMixin, View):
    def get(self, request):
        from timetables.forms import BellScheduleForm, PeriodForm
        return render(request, 'timetables/bell_create.html', {
            'form': BellScheduleForm(),
            'period_form': PeriodForm(),
        })

    def post(self, request):
        from timetables.models import BellSchedule, Period
        from timetables.forms import BellScheduleForm
        form = BellScheduleForm(request.POST)
        if form.is_valid():
            bell = form.save(commit=False)
            bell.owner = request.user
            bell.save()
            messages.success(request, f'Bell schedule "{bell.name}" created!')
            return redirect('timetables:bell-detail', pk=bell.pk)
        return render(request, 'timetables/bell_create.html', {'form': form})



class BellScheduleListView(LoginRequiredMixin, View):
    def get(self, request):
        from timetables.models import BellSchedule
        bells = BellSchedule.objects.filter(owner=request.user).prefetch_related('periods')
        return render(request, 'timetables/bells.html', {
            'bells': bells,
            'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
        })
# class AddPeriodView(LoginRequiredMixin, View):
#     def post(self, request, pk):
#         from timetables.models import BellSchedule
#         from timetables.forms import PeriodForm
#         bell = get_object_or_404(BellSchedule, pk=pk, owner=request.user)
#         form = PeriodForm(request.POST)
#         if form.is_valid():
#             period = form.save(commit=False)
#             period.bell_schedule = bell
#             period.save()
#             messages.success(request, f'Period {period.period_number} added.')
#         else:
#             for f, errs in form.errors.items():
#                 for e in errs:
#                     messages.error(request, f'{f}: {e}')
#         return redirect('timetables:bell-detail', pk=pk)

class AddPeriodView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import BellSchedule
        from timetables.forms import PeriodForm
        bell = get_object_or_404(BellSchedule, pk=pk, owner=request.user)
        return render(request, 'timetables/add_period.html', {
            'bell': bell,
            'form': PeriodForm(),
        })

    def post(self, request, pk):
        from timetables.models import BellSchedule, Period
        from timetables.forms import PeriodForm
        bell = get_object_or_404(BellSchedule, pk=pk, owner=request.user)
        form = PeriodForm(request.POST)
        if form.is_valid():
            period = form.save(commit=False)
            period.bell_schedule = bell
            period.save()
            messages.success(request, 'Period added!')
            return redirect('timetables:bell-detail', pk=bell.pk)
        return render(request, 'timetables/add_period.html', {
            'bell': bell,
            'form': form,
        })
class DeletePeriodView(LoginRequiredMixin, View):
    def post(self, request, pk, period_pk):
        from timetables.models import BellSchedule, Period
        bell = get_object_or_404(BellSchedule, pk=pk, owner=request.user)
        period = get_object_or_404(Period, pk=period_pk, bell_schedule=bell)
        period.delete()
        messages.success(request, 'Period removed.')
        return redirect('timetables:bell-detail', pk=pk)


class DeleteBellScheduleView(LoginRequiredMixin, View):
    def post(self, request, pk):
        from timetables.models import BellSchedule
        bell = get_object_or_404(BellSchedule, pk=pk, owner=request.user)
        bell.delete()
        messages.success(request, 'Bell schedule deleted.')
        return redirect('timetables:bells')


# ────────────────────────────────────────────────────────────
# ADD / DELETE LESSON
# ────────────────────────────────────────────────────────────
class AddLessonView(LoginRequiredMixin, View):
    def post(self, request, pk):
        from timetables.models import Timetable
        from timetables.forms import LessonForm
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        form = LessonForm(request.POST, user=request.user)
        if form.is_valid():
            lesson = form.save(commit=False)
            lesson.timetable = tt
            lesson.save()
            messages.success(request, 'Lesson added.')
        else:
            for f, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f'{f}: {e}')
        return redirect('timetables:setup', pk=pk)


class DeleteLessonView(LoginRequiredMixin, View):
    def post(self, request, pk, lesson_pk):
        from timetables.models import Timetable, Lesson
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        lesson = get_object_or_404(Lesson, pk=lesson_pk, timetable=tt)
        lesson.delete()
        messages.success(request, 'Lesson removed.')
        return redirect('timetables:setup', pk=pk)


# ────────────────────────────────────────────────────────────
# GENERATE
# ────────────────────────────────────────────────────────────
class GenerateTimetableView(LoginRequiredMixin, View):
    def post(self, request, pk):
        from timetables.models import Timetable
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        if tt.status == 'generating':
            messages.warning(request, 'Already generating, please wait.')
            return redirect('timetables:setup', pk=pk)
        tt.slots.all().delete()
        tt.status = 'generating'
        tt.generation_log = 'Queued…'
        tt.save()
        try:
            from scheduling.tasks import generate_timetable_task
            generate_timetable_task.delay(tt.pk)
            messages.success(request, 'Generation started!')
        except Exception:
            try:
                from scheduling.engine import TimetableGenerator
                from timetables.models import TimetableSlot
                gen = TimetableGenerator(tt)
                slots = gen.generate()
                TimetableSlot.objects.bulk_create(slots, batch_size=500)
                tt.status = 'generated'
                tt.generation_log = gen.get_log()
                tt.save()
                messages.success(request, f'Generated! {len(slots)} slots created.')
            except Exception as e2:
                tt.status = 'failed'
                tt.generation_log = str(e2)
                tt.save()
                messages.error(request, f'Generation failed: {e2}')
        return redirect('timetables:setup', pk=pk)


# ────────────────────────────────────────────────────────────
# STATUS / PUBLISH / VIEW
# ────────────────────────────────────────────────────────────
class TimetableStatusView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import Timetable
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        return JsonResponse({
            'status': tt.status,
            'slot_count': tt.slots.count(),
            'log': (tt.generation_log or '')[-600:],
        })


class PublishTimetableView(LoginRequiredMixin, View):
    def post(self, request, pk):
        from timetables.models import Timetable
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        tt.status = 'published'
        tt.save()
        messages.success(request, 'Timetable published!')
        return redirect('timetables:view', pk=pk)


class TimetableGridView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import Timetable, TimetableSlot
        from resources.models import Division
        from faculty.models import Faculty
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        periods = list(tt.bell_schedule.periods.all().order_by('period_number')) if tt.bell_schedule else []
        teaching_periods = [p for p in periods if p.period_type == 'class']
        slots = list(TimetableSlot.objects.filter(timetable=tt)
                     .select_related('lesson__division__grade', 'lesson__subject',
                                     'lesson__faculty', 'room', 'period'))
        day_names = {d: d for d in ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']}
        slot_map = {}
        for s in slots:
            slot_map.setdefault(s.lesson.division_id, {}).setdefault(s.day, {})[s.period_id] = s
        div_ids = list(dict.fromkeys(s.lesson.division_id for s in slots))
        divisions = list(Division.objects.filter(pk__in=div_ids).select_related('grade'))
        faculty_with_email = list(
            Faculty.objects.filter(owner=request.user, status='active')
            .exclude(email='').values('id', 'name', 'email')
        )
        return render(request, 'timetables/view.html', {
            'timetable': tt,
            'periods': teaching_periods,
            'working_days': [(d, d) for d in (tt.bell_schedule.working_days if tt.bell_schedule else [])],
            'day_names': day_names,
            'divisions': divisions,
            'slot_map': slot_map,
            'faculty_with_email': faculty_with_email,
        })

# ────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ────────────────────────────────────────────────────────────
class ExportExcelView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import Timetable, TimetableSlot
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        slots = (TimetableSlot.objects
                 .filter(timetable=tt)
                 .select_related('lesson__division__grade', 'lesson__subject',
                                 'lesson__faculty', 'room', 'period')
                 .order_by('day', 'period__period_number'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tt.name[:31]

        # Header row
        headers = ['Day', 'Period', 'Division', 'Subject', 'Faculty', 'Room']
        header_font = Font(bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row, slot in enumerate(slots, 2):
            ws.cell(row=row, column=1, value=slot.day)
            ws.cell(row=row, column=2, value=slot.period.period_number if slot.period else '')
            ws.cell(row=row, column=3, value=str(slot.lesson.division) if slot.lesson.division else '')
            ws.cell(row=row, column=4, value=slot.lesson.subject.name if slot.lesson.subject else '')
            ws.cell(row=row, column=5, value=slot.lesson.faculty.name if slot.lesson.faculty else '')
            ws.cell(row=row, column=6, value=slot.room.name if slot.room else '')

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"timetable_{tt.name.replace(' ', '_')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

# ────────────────────────────────────────────────────────────
# EXPORT PDF
# ────────────────────────────────────────────────────────────
class ExportPdfView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import Timetable, TimetableSlot
        from django.template.loader import render_to_string
        import subprocess, tempfile, os

        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        slots = (TimetableSlot.objects
                 .filter(timetable=tt)
                 .select_related('lesson__division__grade', 'lesson__subject',
                                 'lesson__faculty', 'room', 'period')
                 .order_by('day', 'period__period_number'))

        html = render_to_string('timetables/export_pdf.html', {
            'timetable': tt,
            'slots': slots,
        }, request=request)

        # Try weasyprint, fall back to plain HTML download
        try:
            from weasyprint import HTML
            pdf = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"timetable_{tt.name.replace(' ', '_')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except ImportError:
            # weasyprint not installed — return HTML as fallback
            return HttpResponse(html, content_type='text/html')


class SendEmailView(LoginRequiredMixin, View):
    def post(self, request, pk):
        import json
        import threading
        from timetables.models import Timetable
        from faculty.models import Faculty

        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)

        try:
            data = json.loads(request.body)
            faculty_ids = data.get('faculty_ids', [])
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid request.'})

        # Validate faculty exist before threading
        faculty_list = list(Faculty.objects.filter(pk__in=faculty_ids, owner=request.user).exclude(email=''))
        if not faculty_list:
            return JsonResponse({'success': False, 'error': 'No valid faculty with emails found.'})

        # Fire and forget in background thread
        def send_all():
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from django.core.mail import EmailMessage
            from django.conf import settings
            from timetables.models import TimetableSlot
            import io

            for faculty in faculty_list:
                try:
                    slots = (TimetableSlot.objects
                             .filter(timetable=tt, lesson__faculty=faculty)
                             .select_related('lesson__subject', 'lesson__division__grade', 'room', 'period')
                             .order_by('day', 'period__period_number'))

                    lines = [f'Timetable: {tt.name}', f'Faculty: {faculty.name}', '']
                    for slot in slots:
                        period_no = slot.period.period_number if slot.period else '?'
                        subject   = slot.lesson.subject.name if slot.lesson.subject else '?'
                        division  = str(slot.lesson.division) if slot.lesson.division else '?'
                        room      = slot.room.name if slot.room else 'TBD'
                        lines.append(f'{slot.day}  |  Period {period_no}  |  {subject}  |  {division}  |  Room: {room}')

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = tt.name[:31]
                    headers = ['Day', 'Period', 'Division', 'Subject', 'Room']
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    for row, slot in enumerate(slots, 2):
                        ws.cell(row=row, column=1, value=slot.day)
                        ws.cell(row=row, column=2, value=slot.period.period_number if slot.period else '')
                        ws.cell(row=row, column=3, value=str(slot.lesson.division) if slot.lesson.division else '')
                        ws.cell(row=row, column=4, value=slot.lesson.subject.name if slot.lesson.subject else '')
                        ws.cell(row=row, column=5, value=slot.room.name if slot.room else '')
                    for col in ws.columns:
                        max_len = max((len(str(c.value or '')) for c in col), default=10)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                    excel_buf = io.BytesIO()
                    wb.save(excel_buf)
                    excel_buf.seek(0)

                    email = EmailMessage(
                        subject=f'Your Timetable — {tt.name}',
                        body='\n'.join(lines),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[faculty.email],
                    )
                    email.attach(
                        f'timetable_{faculty.name.replace(" ", "_")}.xlsx',
                        excel_buf.read(),
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    email.send(fail_silently=True)

                except Exception as e:
                    print(f'Email failed for {faculty.name}: {e}')

        thread = threading.Thread(target=send_all, daemon=True)
        thread.start()

        # Respond immediately
        return JsonResponse({
            'success': True,
            'sent_to': [f.name for f in faculty_list],
            'errors': [],
        })
class TimetableLiveView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from timetables.models import Timetable
        tt = get_object_or_404(Timetable, pk=pk, owner=request.user)
        return render(request, 'timetables/view.html', {
            'timetable': tt,
            'live_mode': True,
        })